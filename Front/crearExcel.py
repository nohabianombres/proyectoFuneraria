import openpyxl
from openpyxl import Workbook
from datetime import datetime
import os

def crear_excel_saldo (datos_antes_de_cambio):
        # Crear un archivo Excel y agregar los datos extraídos
        wb = Workbook()
        ws = wb.active
        ws.title = "Saldos"

        # Agregar encabezados
        encabezados = ['Socio', 'Gasto', 'Descripciones', 'Fecha', 'Valor', 'Funeraria', 'Jefe1', 'Jefe2']
        ws.append(encabezados)

        # Agregar datos extraídos
        for dato in datos_antes_de_cambio:
            fila = []
            for item in dato:
                if item is None:
                    fila.append("")
                elif isinstance(item, list):
                    # Convertir la lista en una cadena separada por comas
                    fila.append(", ".join(map(str, item)))
                else:
                    fila.append(str(item))
            ws.append(fila)

        # Obtener la fecha actual y formatearla
        fecha_actual = datetime.now().strftime("%Y-%m-%d")

        nombre_excel = 'saldos_liquidados_'+fecha_actual+'.xlsx'
        project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
        formularios_path = os.path.join(project_root, 'Front', 'formularios')
        archivo_excel = os.path.join(formularios_path, nombre_excel)
        wb.save(archivo_excel)




def crear_excel_gastos(datos_antes_de_cambio):
    # Crear un archivo Excel y agregar los datos extraídos
    wb = Workbook()
    ws = wb.active
    ws.title = "Gastos"

    # Agregar encabezados
    encabezados = ['Id_gasto', 'Gasto', 'Valor', 'Nombre usuario', 'Fecha']
    ws.append(encabezados)

    # Agregar datos extraídos
    for dato in datos_antes_de_cambio:
        fila = []
        for item in dato:
            if item is None:
                fila.append("")
            elif isinstance(item, list):
                # Convertir la lista en una cadena separada por comas
                fila.append(", ".join(map(str, item)))
            else:
                fila.append(str(item))
        ws.append(fila)

    # Obtener la fecha actual y formatearla
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    nombre_excel = 'gastos_liquidados_'+fecha_actual+'.xlsx'
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    formularios_path = os.path.join(project_root, 'Front', 'formularios')
    archivo_excel = os.path.join(formularios_path, nombre_excel)
    wb.save(archivo_excel)
    print(f"Datos de gastos guardados en el archivo {archivo_excel}")


def crear_excel_colillas(datos_colillas):
    # Crear un archivo Excel y agregar los datos extraídos
    wb = Workbook()
    ws = wb.active
    ws.title = "Colillas"

    # Agregar encabezados
    encabezados = ['Socio', 'Valor mes', 'Desde fecha', 'Hasta fecha', 'Usuario', 'Fecha pago']
    ws.append(encabezados)

    # Agregar datos extraídos
    for dato in datos_colillas:
        fila = []
        for item in dato:
            if item is None:
                fila.append("")
            elif isinstance(item, list):
                # Convertir la lista en una cadena separada por comas
                fila.append(", ".join(map(str, item)))
            else:
                fila.append(str(item))
        ws.append(fila)

    # Obtener la fecha actual y formatearla
    fecha_actual = datetime.now().strftime("%Y-%m-%d")


    nombre_excel = 'colillas_liquidadas_'+fecha_actual+'.xlsx'
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    formularios_path = os.path.join(project_root, 'Front', 'formularios')
    archivo_excel = os.path.join(formularios_path, nombre_excel)
    wb.save(archivo_excel)


def crear_excel_adicionales(datos_facturas):
    # Crear un archivo Excel y agregar los datos extraídos
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas Adicionales"

    # Agregar encabezados
    encabezados = ['Id_factura', 'Nombre comprador', 'Documento comprador', 'Nombre vendedor', 'Valor abonado', 'Fecha']
    ws.append(encabezados)

    # Agregar datos extraídos
    for dato in datos_facturas:
        fila = []
        for item in dato:
            if item is None:
                fila.append("")
            elif isinstance(item, list):
                # Convertir la lista en una cadena separada por comas
                fila.append(", ".join(map(str, item)))
            else:
                fila.append(str(item))
        ws.append(fila)

    # Obtener la fecha actual y formatearla
    fecha_actual = datetime.now().strftime("%Y-%m-%d")

    nombre_excel = 'facturas_adicionales_liquidadas_'+fecha_actual+'.xlsx'
    project_root = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
    formularios_path = os.path.join(project_root, 'Front', 'formularios')
    archivo_excel = os.path.join(formularios_path, nombre_excel)
    wb.save(archivo_excel)